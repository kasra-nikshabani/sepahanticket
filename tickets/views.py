import json
import re
import zipfile
from io import BytesIO
from datetime import timedelta
import logging

import jdatetime
import requests
from django.views.decorators.csrf import csrf_exempt

from django.conf import settings
from django.core.cache import cache
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db import transaction
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponse, JsonResponse, Http404
from django.urls import reverse
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Q
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pytz

import matches
from .models import Ticket, VIPQuota, DiscountCode, Order
from .forms import BulkTicketForm, VIPQuotaForm, DiscountCodeForm
from .reservation import SeatReservation
from matches.models import Match, Seat, Row, MatchSeat, Block, get_block_price_map, get_basa_discount_percent
from accounts.models import User
from wallet.models import Wallet
from .utils import get_access_token

logger = logging.getLogger(__name__)


# ============================================================
#  ویوهای عمومی (کاربران عادی و VIP)
# ============================================================

@login_required
def user_tickets(request):
    match_id = request.GET.get('match_id')
    status_filter = request.GET.get('status')

    if request.user.user_type == 'vip':
        tickets_qs = Ticket.objects.filter(
            user=request.user, status__in=['paid', 'admin_assigned', 'vip_issued', 'cancelled']
        ).order_by('-purchase_date')
    else:
        # 'cancelled' هم نشان داده می‌شود -- وگرنه وقتی ادمین به‌خاطر لغو
        # مسابقه بلیط را باطل و وجه را به کیف پول برمی‌گرداند، بلیط بدون هیچ
        # توضیحی از لیست کاربر ناپدید می‌شود و تنها اطلاع‌رسانیِ این اتفاق
        # همین صفحه است (فعلاً پیامکی برای این مورد ارسال نمی‌شود).
        tickets_qs = Ticket.objects.filter(user=request.user, status__in=['paid', 'cancelled']).order_by('-purchase_date')

    if status_filter and request.user.user_type == 'vip':
        if status_filter in ['paid', 'admin_assigned', 'vip_issued', 'cancelled']:
            tickets_qs = tickets_qs.filter(status=status_filter)

    # فیلتر «مسابقه» باید همه‌ی مسابقاتی که کاربر واقعاً برایشان بلیط دارد را
    # نشان دهد، نه فقط مسابقات فعال -- وگرنه کاربر نمی‌تواند بلیط‌های یک
    # مسابقه‌ی غیرفعال‌شده را فیلتر/پیدا کند (خودِ لیست اصلی بلیط‌ها از قبل
    # بدون قید is_active است و همیشه همه را نشان می‌دهد). باید قبل از اعمال
    # فیلتر match_id محاسبه شود، وگرنه با انتخاب یک مسابقه فقط همان یک
    # گزینه در دراپ‌داون باقی می‌ماند.
    matches = Match.objects.filter(
        id__in=tickets_qs.values_list('match_id', flat=True).distinct()
    ).order_by('-date_time')

    if match_id:
        tickets_qs = tickets_qs.filter(match_id=match_id)

    local_tz = pytz.timezone('Asia/Tehran')
    for ticket in tickets_qs:
        if ticket.used_at:
            ticket.used_at_local = ticket.used_at.astimezone(local_tz)
        else:
            ticket.used_at_local = None

    context = {
        'tickets': tickets_qs,
        'matches': matches,
        'selected_match_id': int(match_id) if match_id else None,
        'selected_status': status_filter,
    }
    return render(request, 'tickets/user_tickets.html', context)


@login_required
def download_ticket_pdf(request, ticket_id):
    """
    دانلود PDF بلیط فقط برای صاحب بلیط یا ادمین.

    فایل‌های media/ticket_pdfs و media/qr_codes روی nginx به‌صورت internal
    تنظیم شده‌اند (دیگه از طریق /media/... مستقیم و بدون احراز هویت در
    دسترس نیستند)؛ این ویو بعد از چک مالکیت، با هدر X-Accel-Redirect از
    nginx می‌خواهد فایل را serve کند.
    """
    ticket = get_object_or_404(Ticket, id=ticket_id)

    is_owner = ticket.user_id == request.user.id
    is_staff = request.user.is_staff or request.user.user_type == 'admin'
    if not (is_owner or is_staff):
        raise Http404

    if not ticket.pdf_file:
        raise Http404

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ticket_{ticket.ticket_number}.pdf"'
    response['X-Accel-Redirect'] = f'/{settings.MEDIA_URL.strip("/")}/{ticket.pdf_file.name}'
    return response


def shared_ticket_pdf(request, token):
    """
    دانلود PDF بلیط از طریق لینک اشتراک‌گذاری -- عمداً بدون نیاز به لاگین،
    چون قرار است هر کسی که خریدار لینک را برایش فرستاده (نه لزوماً خودِ
    خریدار) بتواند بلیط را دانلود کند. امنیتش روی غیرقابل‌حدس بودنِ خودِ
    توکن (UUID4) تکیه دارد، نه احراز هویت -- شبیه لینک‌های اشتراک‌گذاری
    گوگل‌درایو/دراپ‌باکس.
    """
    ticket = get_object_or_404(Ticket, share_token=token)

    if not ticket.pdf_file:
        raise Http404

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ticket_{ticket.ticket_number}.pdf"'
    response['X-Accel-Redirect'] = f'/{settings.MEDIA_URL.strip("/")}/{ticket.pdf_file.name}'
    return response


@login_required
def tickets_pdf_status(request):
    """
    وضعیت آماده بودن PDF چند بلیط را یک‌جا برمی‌گرداند -- برای Polling از
    صفحاتی مثل «بلیط‌های من» که بلافاصله بعد از صدور، PDF هنوز در صف
    پس‌زمینه (RQ) در حال تولید است و کاربر مجبور بود صفحه را رفرش کند تا
    دکمه‌ی دانلود ظاهر شود.
    """
    ids_param = request.GET.get('ids', '')
    try:
        ticket_ids = [int(pk) for pk in ids_param.split(',') if pk.strip()]
    except ValueError:
        return JsonResponse({'error': 'invalid ids'}, status=400)

    is_staff = request.user.is_staff or request.user.user_type == 'admin'
    tickets = Ticket.objects.filter(id__in=ticket_ids)
    if not is_staff:
        tickets = tickets.filter(user_id=request.user.id)

    result = {}
    for ticket in tickets:
        result[str(ticket.id)] = {
            'ready': bool(ticket.pdf_file),
            'pdf_url': reverse('tickets:download_ticket_pdf', args=[ticket.id]) if ticket.pdf_file else None,
            'share_url': (
                f"{request.scheme}://{request.get_host()}{reverse('tickets:shared_ticket_pdf', args=[ticket.share_token])}"
                if ticket.pdf_file else None
            ),
        }
    return JsonResponse(result)


@login_required
def vip_tickets(request):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')

    match_id = request.GET.get('match_id')
    tickets_qs = Ticket.objects.filter(
        user=request.user, is_admin_assigned=True, status='admin_assigned'
    ).order_by('-purchase_date')
    if match_id:
        tickets_qs = tickets_qs.filter(match_id=match_id)

    matches = Match.objects.filter(is_active=True).order_by('-date_time')
    context = {
        'tickets': tickets_qs,
        'matches': matches,
        'selected_match_id': int(match_id) if match_id else None,
    }
    return render(request, 'tickets/vip_tickets.html', context)


@login_required
def vip_dashboard(request):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')

    match_id = request.GET.get('match_id')
    quotas_qs = VIPQuota.objects.filter(user=request.user).select_related('match')
    if match_id:
        quotas_qs = quotas_qs.filter(match_id=match_id)

    matches = Match.objects.filter(is_active=True).order_by('-date_time')
    available_matches = Match.objects.filter(id__in=quotas_qs.values_list('match_id', flat=True)).order_by('-date_time')

    context = {
        'quotas': quotas_qs,
        'matches': available_matches,
        'all_matches': matches,
        'selected_match_id': int(match_id) if match_id else None,
    }
    return render(request, 'tickets/vip_dashboard.html', context)


@login_required
def vip_issued_tickets(request):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')

    match_id = request.GET.get('match_id')
    tickets_qs = Ticket.objects.filter(user=request.user, status='vip_issued').order_by('-purchase_date')

    selected_match = None
    if match_id:
        tickets_qs = tickets_qs.filter(match_id=match_id)
        selected_match = get_object_or_404(Match, id=match_id)

    matches = Match.objects.filter(is_active=True).order_by('-date_time')

    if not tickets_qs.exists() and not match_id:
        messages.info(request, 'هیچ بلیط صادرشده‌ای یافت نشد.')
    elif not tickets_qs.exists() and match_id:
        messages.info(request, f'هیچ بلیطی برای مسابقه "{selected_match}" صادر نشده است.')

    context = {
        'tickets': tickets_qs,
        'matches': matches,
        'selected_match': selected_match,
        'selected_match_id': int(match_id) if match_id else None,
    }
    return render(request, 'tickets/vip_issued_tickets.html', context)


# ============================================================
#  ویوهای مربوط به صدور بلیط توسط کاربر ویژه (VIP)
# ============================================================

@login_required
def vip_issue_manual(request, match_id):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')

    if request.method == 'POST':
        # قفل کردن رکورد مسابقه برای جلوگیری از ثبت همزمان (Race Condition)
        match = get_object_or_404(Match.objects.select_for_update(), id=match_id, is_active=True)
    else:
        match = get_object_or_404(Match, id=match_id, is_active=True)
    quota = get_object_or_404(VIPQuota, user=request.user, match=match)

    if quota.remaining <= 0:
        messages.error(request, 'ظرفیت شما برای این مسابقه تکمیل شده است.')
        return redirect('tickets:vip_dashboard')

    selected_block_id = request.GET.get('block_id')
    selected_row_id = request.GET.get('row_id')

    blocks = Block.objects.filter(stadium=match.stadium, is_active=True).order_by('order')
    selected_block = None
    rows = []
    all_seats = []

    if selected_block_id:
        selected_block = get_object_or_404(Block, id=selected_block_id)
        rows = Row.objects.filter(block=selected_block, is_active=True).order_by('number')

        if selected_row_id:
            if not rows.filter(id=selected_row_id).exists():
                selected_row_id = None

        seats_qs = Seat.objects.filter(row__block=selected_block, row__is_active=True)
        if selected_row_id:
            seats_qs = seats_qs.filter(row_id=selected_row_id)
        seats_qs = seats_qs.order_by('row__number', 'number')

        for seat in seats_qs:
            match_seat, created = MatchSeat.objects.get_or_create(match=match, seat=seat,
                                                                  defaults={'is_available': True})
            seat.match_seat = match_seat
            seat.is_available = match_seat.is_available
            all_seats.append(seat)

    if request.method == 'POST':
        match_seat_id = request.POST.get('match_seat_id')
        full_name = request.POST.get('full_name')
        national_code = request.POST.get('national_code')

        if not match_seat_id or not full_name or not national_code:
            messages.error(request, 'لطفاً تمام اطلاعات را وارد کنید.')
            return redirect('tickets:vip_issue_manual', match_id=match_id)

        if len(national_code) != 10 or not national_code.isdigit():
            messages.error(request, 'کد ملی باید ۱۰ رقم باشد.')
            return redirect('tickets:vip_issue_manual', match_id=match_id)

        with transaction.atomic():
            # quota بالاتر بدون قفل خونده شده بود؛ دو درخواست هم‌زمان (دو تب یا
            # دابل‌کلیک) می‌تونستن هر دو remaining>0 رو ببینن و هر دو صادر کنن.
            quota = VIPQuota.objects.select_for_update().get(pk=quota.pk)
            if quota.remaining <= 0:
                messages.error(request, 'ظرفیت شما برای این مسابقه تکمیل شده است.')
                return redirect('tickets:vip_dashboard')

            try:
                match_seat = MatchSeat.objects.select_for_update().get(id=match_seat_id, match=match, is_available=True)
            except MatchSeat.DoesNotExist:
                messages.error(request, 'صندلی انتخاب‌شده معتبر نیست یا قبلاً فروخته شده است.')
                return redirect('tickets:vip_issue_manual', match_id=match_id)

            ticket = Ticket.objects.create(
                user=request.user, match=match, seat=match_seat.seat, match_seat=match_seat,
                full_name=full_name, national_code=national_code, status='vip_issued', is_admin_assigned=False,
            )

            match_seat.is_available = False
            match_seat.save()

            quota.used += 1
            quota.save()

        messages.success(request,
                         f'✅ بلیط برای {full_name} با موفقیت صادر شد.\nبلوک: {match_seat.seat.row.block.name} - ردیف: {match_seat.seat.row.number} - صندلی: {match_seat.seat.number}')
        return redirect(f'{reverse("tickets:vip_issued_tickets")}?match_id={match.id}')

    context = {
        'match': match, 'quota': quota, 'blocks': blocks, 'selected_block': selected_block,
        'rows': rows, 'selected_row_id': int(selected_row_id) if selected_row_id else None,
        'all_seats': all_seats,
    }
    return render(request, 'tickets/vip_issue_manual.html', context)


@login_required
def vip_edit_issued_ticket(request, ticket_id):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')

    ticket = get_object_or_404(Ticket, id=ticket_id, user=request.user, status='vip_issued')

    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        national_code = request.POST.get('national_code')

        if not full_name or not national_code:
            messages.error(request, 'لطفاً تمام اطلاعات را وارد کنید.')
            return redirect('tickets:vip_edit_issued_ticket', ticket_id=ticket.id)

        if len(national_code) != 10 or not national_code.isdigit():
            messages.error(request, 'کد ملی باید ۱۰ رقم باشد.')
            return redirect('tickets:vip_edit_issued_ticket', ticket_id=ticket.id)

        ticket.full_name = full_name
        ticket.national_code = national_code
        if ticket.pdf_file:
            ticket.pdf_file.delete(save=False)
            ticket.pdf_file = None
        ticket.save()  # pdf_file خالیه -> تولید مجدد PDF خودکار توی صف پس‌زمینه enqueue می‌شه

        messages.success(request, f'✅ اطلاعات بلیط {ticket.ticket_number} با موفقیت ویرایش شد. تولید PDF جدید چند لحظه طول می‌کشد.')
        return redirect('tickets:vip_issued_tickets')

    return render(request, 'tickets/vip_edit_issued_ticket.html', {'ticket': ticket})


@login_required
def vip_delete_issued_ticket(request, ticket_id):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')

    ticket = get_object_or_404(Ticket, id=ticket_id, user=request.user, status='vip_issued')

    if request.method == 'POST':
        with transaction.atomic():
            if ticket.match_seat:
                ticket.match_seat.is_available = True
                ticket.match_seat.save()

            if ticket.pdf_file:
                ticket.pdf_file.delete(save=False)
            if ticket.qr_code:
                ticket.qr_code.delete(save=False)

            ticket.delete()

        messages.success(request, '✅ بلیط با موفقیت لغو شد و صندلی آزاد شد.')
        return redirect('tickets:vip_issued_tickets')

    return render(request, 'tickets/vip_delete_issued_ticket.html', {'ticket': ticket})


@login_required
def vip_issue_excel(request, match_id):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')

    match = get_object_or_404(Match, id=match_id, is_active=True)
    quota = get_object_or_404(VIPQuota, user=request.user, match=match)

    if quota.remaining <= 0:
        messages.error(request, 'ظرفیت شما برای این مسابقه تکمیل شده است.')
        return redirect('tickets:vip_dashboard')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            import pandas as pd
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            required_columns = ['نام و نام خانوادگی', 'کد ملی']
            if not all(col in df.columns for col in required_columns):
                messages.error(request, 'فایل باید شامل ستون‌های "نام و نام خانوادگی" و "کد ملی" باشد.')
                return redirect('tickets:vip_issue_excel', match_id=match_id)

            with transaction.atomic():
                # quota و لیست صندلی‌های موجود بالاتر بدون قفل خونده شده بودن؛
                # دو آپلود اکسل هم‌زمان می‌تونستن هم روی سهمیه هم روی همون
                # صندلی‌های «موجود» با هم تداخل کنن. اینجا با قفل دوباره چک می‌شه.
                quota = VIPQuota.objects.select_for_update().get(pk=quota.pk)

                available_match_seats = list(
                    MatchSeat.objects.select_for_update().filter(match=match, is_available=True)
                    .select_related('seat')
                    .order_by('seat__row__block__order', 'seat__row__number', 'seat__number')
                )

                if len(available_match_seats) < len(df):
                    messages.error(request,
                                   f'تعداد صندلی‌های موجود ({len(available_match_seats)}) کمتر از تعداد درخواست‌ها ({len(df)}) است.')
                    return redirect('tickets:vip_issue_excel', match_id=match_id)

                if len(df) > quota.remaining:
                    messages.error(request,
                                   f'ظرفیت باقی‌مانده شما ({quota.remaining}) کمتر از تعداد درخواست‌ها ({len(df)}) است.')
                    return redirect('tickets:vip_issue_excel', match_id=match_id)

                created = 0
                for _, row_data in df.iterrows():
                    full_name = str(row_data['نام و نام خانوادگی']).strip()
                    national_code = str(row_data['کد ملی']).strip()

                    if not full_name or not national_code or len(national_code) != 10:
                        continue

                    match_seat = available_match_seats[created]

                    ticket = Ticket.objects.create(
                        user=request.user, match=match, seat=match_seat.seat, match_seat=match_seat,
                        full_name=full_name, national_code=national_code, status='vip_issued', is_admin_assigned=False,
                    )
                    # .create() قبلاً خودش save() را صدا زده -- یک save() اضافه اینجا
                    # چون pdf_file هنوز خالیه (تولید PDF async و صف‌بندی‌شده است)،
                    # باعث می‌شد Ticket.save() دوباره enqueue_pdf_generation را صدا بزند
                    # و همان بلیط دوبار در صف PDF قرار بگیرد.

                    match_seat.is_available = False
                    match_seat.save()
                    created += 1

                quota.used += created
                quota.save()

            messages.success(request, f'✅ {created} بلیط با موفقیت از فایل اکسل صادر شد.')
            return redirect(f'{reverse("tickets:vip_issued_tickets")}?match_id={match.id}')

        except Exception as e:
            messages.error(request, f'❌ خطا در خواندن فایل: {str(e)}')
            return redirect('tickets:vip_issue_excel', match_id=match_id)

    return render(request, 'tickets/vip_issue_excel.html', {'match': match, 'quota': quota})


# ============================================================
#  ویوهای مربوط به خرید (کاربران عادی) – با Redis
# ============================================================

@login_required
def select_seats(request, match_id):
    match = get_object_or_404(Match, id=match_id, is_active=True)
    if not match.ticket_sales_enabled:
        messages.error(request, 'فروش بلیط برای این مسابقه غیرفعال است.')
        return redirect('matches:home')
    row_id = request.session.get('selected_row_id')
    if not row_id:
        messages.error(request, 'لطفاً ابتدا یک ردیف را انتخاب کنید.')
        return redirect('matches:select_block', match_id=match_id)
    if request.user.user_type == 'vip':
        messages.error(request, 'کاربران ویژه امکان خرید بلیط ندارند.')
        return redirect('matches:home')

    section_id = request.session.get('section_id')
    if not section_id:
        messages.error(request, 'لطفاً ابتدا ردیف را انتخاب کنید.')
        return redirect('matches:block_map', match_id=match_id)

    row = get_object_or_404(Row, id=section_id, is_active=True)

    if request.session.get('selected_seats'):
        old_seats = request.session.get('selected_seats', [])
        with transaction.atomic():
            for seat_id in old_seats:
                try:
                    ms = MatchSeat.objects.get(id=seat_id, match=match)
                    ms.is_available = True
                    ms.reserved_until = None
                    ms.save()
                except MatchSeat.DoesNotExist:
                    pass
                SeatReservation.release(seat_id, user_id=getattr(request.user, 'id', None), force=True)
        request.session.pop('selected_seats', None)
        request.session.pop('match_id', None)
        request.session.pop('section_id', None)
        request.session.pop('reserved_at', None)
        messages.info(request, 'رزروهای قبلی شما به دلیل بازگشت به این صفحه لغو شد.')

    match_seats = MatchSeat.objects.filter(match=match, seat__row=row).select_related('seat', 'seat__row').order_by(
        'seat__number')
    if not match_seats.exists():
        messages.warning(request, 'هیچ صندلی‌ای برای این ردیف در این مسابقه تعریف نشده است.')
        return redirect('matches:block_map', match_id=match_id)

    request.session['match_id'] = match_id
    request.session['row_id'] = row.id

    return render(request, 'matches/select_seats.html', {'match': match, 'row': row, 'match_seats': match_seats})


@login_required
def reserve_seats(request, match_id):
    """
    رزرو صندلی با قفل دیتابیس (select_for_update) + Redis atomic (SET NX).
    از oversell و race condition در فشار بالا جلوگیری می‌کند.
    """
    if request.user.user_type == 'vip':
        messages.error(request, 'کاربران ویژه امکان خرید بلیط ندارند.')
        return redirect('matches:home')

    match = get_object_or_404(Match, id=match_id, is_active=True)
    if not match.ticket_sales_enabled:
        messages.error(request, 'فروش بلیط برای این مسابقه غیرفعال است.')
        return redirect('matches:home')

    row_id = request.POST.get('row_id') or request.session.get('selected_row_id')
    if not row_id:
        messages.error(request, 'لطفاً ابتدا یک ردیف را انتخاب کنید.')
        return redirect('matches:select_block', match_id=match_id)

    if request.method != 'POST':
        return redirect('matches:block_map', match_id=match_id)

    selected_seats_data = request.POST.get('selected_seats')
    if not selected_seats_data:
        messages.error(request, 'هیچ صندلی‌ای انتخاب نشده است.')
        return redirect('matches:block_map', match_id=match_id)

    try:
        selected_seats = json.loads(selected_seats_data)
        selected_seats = [int(s) for s in selected_seats]
        if len(selected_seats) > 5:
            messages.error(request, 'حداکثر ۵ بلیط مجاز است.')
            return redirect('matches:block_map', match_id=match_id)
        if len(selected_seats) == 0:
            messages.error(request, 'هیچ صندلی‌ای انتخاب نشده است.')
            return redirect('matches:block_map', match_id=match_id)
    except (json.JSONDecodeError, TypeError, ValueError):
        messages.error(request, 'خطا در اطلاعات ارسال شده.')
        return redirect('matches:block_map', match_id=match_id)

    reserved_seats = []
    timeout = getattr(settings, 'SEAT_RESERVATION_TIMEOUT', 600)
    now = timezone.now()

    try:
        with transaction.atomic():
            # قفل ردیف‌های MatchSeat تا پایان تراکنش — جلوگیری از double-booking
            locked = (
                MatchSeat.objects
                .select_for_update()
                .select_related('seat')
                .filter(id__in=selected_seats, match=match)
            )
            locked_map = {ms.id: ms for ms in locked}

            if len(locked_map) != len(set(selected_seats)):
                messages.error(request, 'یکی از صندلی‌های انتخاب‌شده معتبر نیست.')
                return redirect('matches:block_map', match_id=match_id)

            # صندلی‌ای که خودش (سطح Seat، نه فقط MatchSeat این مسابقه) غیرفعال
            # شده هیچ‌وقت قابل رزرو نیست -- حتی اگر از یک لینک/درخواست قدیمی باشد
            if any(not ms.seat.is_available for ms in locked_map.values()):
                messages.error(request, 'یکی از صندلی‌های انتخاب‌شده در دسترس نیست.')
                return redirect('matches:block_map', match_id=match_id)

            for seat_id in selected_seats:
                match_seat = locked_map[seat_id]

                # اگر قبلاً توسط همین کاربر رزرو شده، تمدید کن
                if SeatReservation.is_reserved_by_user(seat_id, request.user.id):
                    SeatReservation.extend_reservation(seat_id, user_id=request.user.id)
                    match_seat.reserved_until = now + timedelta(seconds=timeout)
                    match_seat.is_available = False
                    match_seat.save(update_fields=['is_available', 'reserved_until'])
                    reserved_seats.append(seat_id)
                    continue

                # صندلی فروخته‌شده یا رزرو فعال دیگران
                if not match_seat.is_available:
                    expired = (
                        match_seat.reserved_until is not None
                        and match_seat.reserved_until < now
                    )
                    if not expired:
                        SeatReservation.release_many(reserved_seats, user_id=getattr(request.user, 'id', None), force=True)
                        messages.error(
                            request,
                            'یکی از صندلی‌های انتخاب‌شده قبلاً فروخته یا رزرو شده است.'
                        )
                        return redirect('matches:block_map', match_id=match_id)
                    # رزرو منقضی شده — آزادسازی و ادامه

                success, msg = SeatReservation.reserve(seat_id, request.user.id, match_id)
                if not success:
                    SeatReservation.release_many(reserved_seats, user_id=getattr(request.user, 'id', None), force=True)
                    messages.error(request, msg)
                    return redirect('matches:block_map', match_id=match_id)

                match_seat.is_available = False
                match_seat.reserved_until = now + timedelta(seconds=timeout)
                match_seat.save(update_fields=['is_available', 'reserved_until'])
                reserved_seats.append(seat_id)

    except Exception as e:
        SeatReservation.release_many(reserved_seats, user_id=getattr(request.user, 'id', None), force=True)
        logger.error(f"Error in reserve_seats: {str(e)}", exc_info=True)
        messages.error(request, 'خطا در رزرو صندلی. لطفاً دوباره تلاش کنید.')
        return redirect('matches:block_map', match_id=match_id)

    request.session['selected_seats'] = selected_seats
    request.session['match_id'] = match_id
    request.session['section_id'] = row_id
    request.session['reserved_at'] = now.isoformat()

    return redirect('tickets:ticket_info', match_id=match_id)


# ============================================================
#  تابع محاسبه سن
# ============================================================
def get_age_from_jalali(dob_str):
    try:
        parts = dob_str.split('/')
        if len(parts) == 3:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            birth_date = jdatetime.date(y, m, d)
            today = jdatetime.date.today()
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
            return age
        return 100
    except Exception:
        return 100


# ============================================================
#  ویو اطلاعات خریدار و پرداخت
# ============================================================
@never_cache
@login_required
def ticket_info(request, match_id):
    if request.user.user_type == 'vip':
        messages.error(request, 'کاربران ویژه امکان خرید بلیط ندارند.')
        return redirect('matches:home')

    match = get_object_or_404(Match, id=match_id, is_active=True)
    if not match.ticket_sales_enabled:
        messages.error(request, 'فروش بلیط برای این مسابقه غیرفعال است.')
        return redirect('matches:home')
    selected_seats = request.session.get('selected_seats', [])

    if not selected_seats:
        messages.error(request, 'ابتدا صندلی‌ها را انتخاب کنید.')
        return redirect('matches:block_map', match_id=match_id)

    for seat_id in selected_seats:
        redis_ok = SeatReservation.is_reserved(seat_id)
        if not redis_ok:
            try:
                ms = MatchSeat.objects.get(id=seat_id, match=match)
                db_ok = (not ms.is_available) and (ms.reserved_until is None or ms.reserved_until > timezone.now())
            except MatchSeat.DoesNotExist:
                db_ok = False

            if not db_ok:
                with transaction.atomic():
                    for sid in selected_seats:
                        try:
                            ms2 = MatchSeat.objects.get(id=sid, match=match)
                            ms2.is_available = True
                            ms2.reserved_until = None
                            ms2.save()
                        except MatchSeat.DoesNotExist:
                            pass
                request.session.pop('selected_seats', None)
                request.session.pop('reserved_at', None)
                messages.error(request, 'رزرو یکی از صندلی‌ها منقضی شده است. لطفاً دوباره انتخاب کنید.')
                return redirect('matches:block_map', match_id=match_id)

    reserved_at = request.session.get('reserved_at')
    remaining_time = 0
    if reserved_at:
        from datetime import datetime
        reserved_time = datetime.fromisoformat(reserved_at)
        elapsed = (timezone.now() - reserved_time).total_seconds()
        remaining_time = max(0, int(settings.SEAT_RESERVATION_TIMEOUT - elapsed))

    if remaining_time <= 0 and reserved_at:
        return redirect('tickets:cancel_reservation', match_id=match_id)

    block_price_map = get_block_price_map(match)

    seats_data = []
    total_price = 0
    for seat_id in selected_seats:
        try:
            match_seat = MatchSeat.objects.select_related('seat__row__block').get(id=seat_id, match=match)
            block = match_seat.seat.row.block
            block_price = block_price_map.get(block.id, block.price)
            price = int(block_price) if block_price else 0
            seats_data.append({
                'id': match_seat.id,
                'number': match_seat.seat.number,
                'row_name': match_seat.seat.row.name,
                'is_home': match_seat.seat.row.is_home,
                'price': price,
                'zone_label': match_seat.seat.row.zone_label,
            })
            total_price += price
        except MatchSeat.DoesNotExist:
            messages.error(request, 'یکی از صندلی‌های انتخاب‌شده معتبر نیست.')
            return redirect('matches:block_map', match_id=match_id)

    wallet, created = Wallet.objects.get_or_create(user=request.user)

    from accounts.models import SiteSettings
    context = {
        'match': match,
        'seats_data': seats_data,
        'total_price': total_price,
        'remaining_time': remaining_time,
        'wallet_balance': wallet.balance,
        'bypass_civil_registry_inquiry': SiteSettings.get_solo().bypass_civil_registry_inquiry,
    }

    if request.method == 'POST':
        use_wallet = request.POST.get('use_wallet') == 'true'
        try:
            wallet_amount = int(request.POST.get('wallet_amount', 0))
        except (ValueError, TypeError):
            wallet_amount = 0

        discount_code = request.POST.get('discount_code', '').strip()
        discount_obj = None
        discount_percent = 0

        if discount_code:
            try:
                discount_obj = DiscountCode.objects.get(code=discount_code)
                valid, msg = discount_obj.is_valid(match=match)
                if not valid:
                    messages.error(request, f'کد تخفیف نامعتبر: {msg}')
                    return render(request, 'tickets/ticket_info.html', context)
                discount_percent = discount_obj.discount_percent
            except DiscountCode.DoesNotExist:
                messages.error(request, 'کد تخفیف یافت نشد.')
                return render(request, 'tickets/ticket_info.html', context)

        actual_total_price = 0
        processed_tickets_data = []
        basa_discount_percent = get_basa_discount_percent(match)
        basa_national_codes = set()
        if basa_discount_percent > 0:
            basa_national_codes = set(User.objects.filter(is_basa_member=True).values_list('national_code', flat=True))

        for seat_data in seats_data:
            match_seat_id = seat_data['id']
            full_name = request.POST.get(f'full_name_{match_seat_id}')

            # اگر کاربر این صندلی را در صفحه حذف کرده باشد، از پردازش رد می‌شود
            if not full_name:
                continue

            national_code = request.POST.get(f'national_code_{match_seat_id}')
            tarikhe_tavallod = request.POST.get(f'tarikhe_tavallod_{match_seat_id}')
            shomare_hamrah = request.POST.get(f'shomare_hamrah_{match_seat_id}')

            if not all([full_name, national_code, tarikhe_tavallod, shomare_hamrah]):
                messages.error(request, 'لطفاً تمام اطلاعات را برای همه صندلی‌ها وارد و استعلام کنید.')
                return render(request, 'tickets/ticket_info.html', context)

            if len(national_code) != 10 or not national_code.isdigit():
                messages.error(request, f'کد ملی برای صندلی {seat_data["number"]} باید ۱۰ رقم باشد.')
                return render(request, 'tickets/ticket_info.html', context)

            # ===== بررسی محدودیت خرید (VIP و ۱ بلیط در هر مسابقه) =====
            if request.user.user_type != 'vip':
                if national_code in [t['national_code'] for t in processed_tickets_data]:
                    messages.error(request,
                                   f'کد ملی {national_code} برای دو صندلی مختلف وارد شده است. هر فرد فقط یک بلیط می‌تواند داشته باشد.')
                    return render(request, 'tickets/ticket_info.html', context)

                has_ticket = Ticket.objects.filter(
                    national_code=national_code, status='paid', match=match
                ).exists()
                if has_ticket:
                    messages.error(request, f'کد ملی {national_code} قبلاً برای این مسابقه بلیط خریداری کرده است.')
                    return render(request, 'tickets/ticket_info.html', context)

            age = get_age_from_jalali(tarikhe_tavallod)
            basa_discount_amount = 0
            if age < 15:
                seat_price = 0
            else:
                seat_price = seat_data['price']
                # ===== تخفیف باسا -- فقط روی همین بلیط، نه کل سبد خرید =====
                # اگر کد ملیِ همین بلیط عضو باسا باشد و برای این مسابقه تخفیف
                # باسا تعیین شده باشد، فقط قیمت همین بلیط تخفیف می‌خورد؛ سایر
                # بلیط‌های همین سفارش (برای کدملی‌های دیگر) دست‌نخورده می‌مانند.
                if basa_discount_percent > 0 and national_code in basa_national_codes:
                    discounted_price = seat_price - int(seat_price * basa_discount_percent / 100)
                    basa_discount_amount = seat_price - discounted_price
                    seat_price = discounted_price

            actual_total_price += seat_price
            processed_tickets_data.append({
                'match_seat_id': match_seat_id,
                'full_name': full_name,
                'national_code': national_code,
                'price': seat_price,
                'age': age,
                'basa_discount_amount': basa_discount_amount,
            })

        if not processed_tickets_data:
            messages.error(request, 'شما هیچ صندلی‌ای برای خرید انتخاب نکرده‌اید.')
            return render(request, 'tickets/ticket_info.html', context)

        discounted_total = actual_total_price - int(actual_total_price * discount_percent / 100)

        wallet_amount_used = 0
        if use_wallet and wallet_amount > 0:
            if wallet_amount > wallet.balance:
                messages.error(request, f'موجودی کیف پول شما ({wallet.balance:,} ریال) کافی نیست.')
                return render(request, 'tickets/ticket_info.html', context)
            if wallet_amount > discounted_total:
                messages.error(request,
                               f'مبلغ قابل پرداخت از کیف پول نمی‌تواند از مبلغ نهایی ({discounted_total:,} ریال) بیشتر باشد.')
                return render(request, 'tickets/ticket_info.html', context)
            wallet_amount_used = wallet_amount

        final_amount = discounted_total - wallet_amount_used

        if final_amount > 0:
            messages.error(request, 'مبلغ نهایی خرید بزرگ‌تر از صفر است. لطفاً از طریق دکمه‌ی پرداخت اقدام کنید.')
            return render(request, 'tickets/ticket_info.html', context)

        try:
            with transaction.atomic():
                # ===== قفل و اعتبارسنجی مجدد کد تخفیف (جلوگیری از race condition) =====
                # discount_obj بالاتر بدون قفل خونده و چک شده بود؛ دو خرید همزمان
                # با همون کد می‌تونستن هر دو رد شرط used_count < max_uses رو
                # ببینن و هر دو used_count رو افزایش بدن و از max_uses رد بشه.
                if discount_obj:
                    discount_obj = DiscountCode.objects.select_for_update().get(pk=discount_obj.pk)
                    valid, msg = discount_obj.is_valid(match=match)
                    if not valid:
                        messages.error(request, f'کد تخفیف دیگر معتبر نیست: {msg}')
                        return render(request, 'tickets/ticket_info.html', context)

                payment_method = 'free'
                if wallet_amount_used > 0 and discount_percent > 0:
                    payment_method = 'mixed'
                elif wallet_amount_used > 0:
                    payment_method = 'wallet'

                order = Order.objects.create(
                    user=request.user, match=match, subtotal=actual_total_price,
                    discount_percent=discount_percent,
                    discount_amount=int(actual_total_price * discount_percent / 100),
                    # مبلغ نهایی سفارش (پس از تخفیف) -- قبلاً همیشه ۰ ثبت می‌شد چون این
                    # مسیر خرید (کاملاً از کیف پول یا رایگان، بدون درگاه) از الگوی
                    # دیگرِ خرید (payments/views.py) که total_amount را درست محاسبه
                    # می‌کند پیروی نمی‌کرد. این باعث می‌شد جاهایی که برای بازگشت وجه یا
                    # گزارش‌گیری به Order.total_amount اعتماد می‌کنند (مثل لغو مسابقه)،
                    # برای بلیط‌های خریداری‌شده‌ی کاملاً با کیف پول مبلغ صفر ببینند.
                    total_amount=discounted_total, wallet_balance_before=wallet.balance,
                    wallet_amount=wallet_amount_used, wallet_balance_after=wallet.balance - wallet_amount_used,
                    payment_method=payment_method, payment_status='paid',
                    discount_code=discount_obj.code if discount_obj else '',
                    discount_code_id=discount_obj.id if discount_obj else None,
                    full_name=request.user.get_full_name() or request.user.username,
                    phone_number=getattr(request.user, 'phone_number', ''),
                    paid_at=timezone.now(),
                )

                tickets = []
                for ticket_data in processed_tickets_data:
                    match_seat_id = ticket_data['match_seat_id']

                    try:
                        match_seat = MatchSeat.objects.get(id=match_seat_id, match=match)
                    except MatchSeat.DoesNotExist:
                        messages.error(request, 'صندلی مورد نظر معتبر نیست یا قبلاً فروخته شده است.')
                        return redirect('matches:block_map', match_id=match_id)

                    if match_seat.reserved_until and match_seat.reserved_until < timezone.now():
                        match_seat.is_available = True
                        match_seat.reserved_until = None
                        match_seat.save()
                        SeatReservation.release(match_seat_id, user_id=getattr(request.user, 'id', None), force=True)
                        messages.error(request, f'مدت زمان رزرو صندلی {match_seat.seat.number} به پایان رسیده است.')
                        return redirect('matches:block_map', match_id=match_id)

                    ticket = Ticket.objects.create(
                        user=request.user, match=match, seat=match_seat.seat, match_seat=match_seat,
                        full_name=ticket_data['full_name'], national_code=ticket_data['national_code'],
                        status='paid', is_admin_assigned=False, order=order,
                        price=ticket_data['price'],  # ذخیره قیمت نهایی (0 برای زیر 15 سال)
                        age=ticket_data['age'],
                        basa_discount_amount=ticket_data['basa_discount_amount'],
                    )
                    tickets.append(ticket)

                    match_seat.is_available = False
                    match_seat.reserved_until = None
                    match_seat.save()
                    SeatReservation.release(match_seat_id, user_id=getattr(request.user, 'id', None), force=True)

                if wallet_amount_used > 0:
                    success = wallet.deduct_balance(
                        amount=wallet_amount_used,
                        description=f"پرداخت {wallet_amount_used:,} ریال از کل {discounted_total:,} ریال (با تخفیف {discount_percent}%) برای خرید {len(tickets)} بلیط - مسابقه {match.home_team} vs {match.away_team}",
                        reference_id=order.order_number, tx_type='ticket_purchase'
                    )
                    if not success:
                        raise Exception("کسر از کیف پول ناموفق بود")
                    wallet.refresh_from_db()
                    order.wallet_balance_after = wallet.balance
                    order.save(update_fields=['wallet_balance_after'])

                if discount_obj:
                    discount_obj.used_count += 1
                    discount_obj.save()

                request.session.pop('selected_seats', None)
                request.session.pop('match_id', None)
                request.session.pop('section_id', None)
                request.session.pop('reserved_at', None)

                if wallet_amount_used > 0:
                    messages.success(request,
                                     f'✅ {len(tickets)} بلیط با موفقیت صادر شد. مبلغ {wallet_amount_used:,} ریال از کیف پول کسر شد.')
                else:
                    messages.success(request, f'✅ {len(tickets)} بلیط به صورت رایگان صادر شد.')

                return redirect('tickets:user_tickets')

        except Exception as e:
            logger.error(f"❌ خطا در پردازش خرید: {str(e)}")
            messages.error(request, f'خطا در پردازش خرید: {str(e)}')
            return render(request, 'tickets/ticket_info.html', context)

    return render(request, 'tickets/ticket_info.html', context)


@login_required
def cancel_reservation(request, match_id):
    match = get_object_or_404(Match, id=match_id, is_active=True)
    selected_seats = request.session.get('selected_seats', [])

    if selected_seats:
        for seat_id in selected_seats:
            try:
                match_seat = MatchSeat.objects.get(id=seat_id, match=match)
                match_seat.is_available = True
                match_seat.reserved_until = None
                match_seat.save(update_fields=['is_available', 'reserved_until'])
            except MatchSeat.DoesNotExist:
                pass
            SeatReservation.release(seat_id, user_id=getattr(request.user, 'id', None), force=True)

        request.session.pop('selected_seats', None)
        request.session.pop('match_id', None)
        request.session.pop('section_id', None)
        request.session.pop('reserved_at', None)

        messages.info(request, 'عملیات خرید لغو شد و صندلی‌ها آزاد شدند.')

    return redirect('matches:select_block', match_id=match_id)


@login_required
@csrf_exempt
def release_reservation(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    match_id = request.POST.get('match_id')
    if not match_id:
        return JsonResponse({'status': 'error', 'message': 'match_id required'}, status=400)

    try:
        match = Match.objects.get(id=match_id)
    except Match.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Match not found'}, status=404)

    selected_seats = request.session.get('selected_seats', [])
    if not selected_seats:
        return JsonResponse({'status': 'ok', 'message': 'No active reservation'})

    with transaction.atomic():
        for seat_id in selected_seats:
            try:
                match_seat = MatchSeat.objects.get(id=seat_id, match=match)
                has_paid_ticket = Ticket.objects.filter(match_seat=match_seat, status='paid').exists()
                if not has_paid_ticket:
                    match_seat.is_available = True
                    match_seat.reserved_until = None
                    match_seat.save()
                    SeatReservation.release(seat_id, user_id=getattr(request.user, 'id', None), force=True)
            except MatchSeat.DoesNotExist:
                pass

    request.session.pop('selected_seats', None)
    request.session.pop('match_id', None)
    request.session.pop('section_id', None)
    request.session.pop('reserved_at', None)

    return JsonResponse({'status': 'ok', 'message': 'Reservation released'})


PERSIAN_NAME_RE = re.compile(r'^[؀-ۿ‌ ]+$')


# ============================================================
#  استعلام از ثبت احوال
# ============================================================
@never_cache
@login_required
def inquiry_fan(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # ===== همه‌چیز (حتی چک نرخ‌محدودیت) داخل try است تا این ویو هرگز صفحه‌ی
    # خطای HTML جنگو/سرور را برنگرداند -- فرانت‌اند همیشه response.json() صدا
    # می‌زند و اگر یک خطای پیش‌بینی‌نشده (مثلاً یک قطعی لحظه‌ای Redis) این
    # try را دور بزند، به‌جای JSON یک صفحه‌ی HTML برمی‌گردد و همان خطای
    # "Unexpected token '<'" سمت کاربر ظاهر می‌شود =====
    try:
        # ===== محدودیت تعداد استعلام (جلوگیری از سوءاستفاده از API ثبت‌احوال) =====
        rate_key = f'inquiry_fan_count:{request.user.id}'
        count = cache.get(rate_key, 0)
        if count >= 20:
            return JsonResponse(
                {'success': False, 'error': 'تعداد استعلام‌های شما زیاد است. چند دقیقه صبر کنید.'}, status=429
            )
        cache.set(rate_key, count + 1, timeout=600)  # ۱۰ دقیقه

        name = request.POST.get('name')
        famil = request.POST.get('famil')
        kode_meli = request.POST.get('kodeMeli')
        shomare_hamrah = request.POST.get('shomareyeHamrah')
        tarikhe_tavallod = request.POST.get('tarikheTavallod')
        match_id = request.POST.get('match_id')

        if not all([name, famil, kode_meli, shomare_hamrah, tarikhe_tavallod]):
            return JsonResponse({'success': False, 'error': 'تمام فیلدها الزامی است'}, status=400)

        if not PERSIAN_NAME_RE.match(name.strip()) or not PERSIAN_NAME_RE.match(famil.strip()):
            return JsonResponse(
                {'success': False, 'error': 'نام و نام خانوادگی باید فقط با حروف فارسی نوشته شود.'}, status=400
            )

        if len(kode_meli) != 10:
            return JsonResponse({'success': False, 'error': 'کد ملی باید ۱۰ رقم باشد'}, status=400)

        token = get_access_token()
        if not token:
            return JsonResponse({'success': False, 'error': 'خطا در دریافت توکن'}, status=401)

        url = "https://fans.footballeticket.ir/api/v1/user/register"  # یا آدرس API اصلی که به شما دادند
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {token}'
        }

        payload = {
            "name": name,
            "famil": famil,
            "kodeMeli": kode_meli,
            "shomareHamrah": shomare_hamrah,
            "tarikheTavallod": tarikhe_tavallod
        }

        response = requests.post(url, json=payload, headers=headers, timeout=10)

        # ===== بررسی اینکه آیا سرور پاسخ JSON داده یا خیر =====
        try:
            data = response.json()
        except ValueError:
            # اگر سرور JSON برنگرداند (مثلا صفحه ارور بدهد)
            return JsonResponse({
                'success': False,
                'error': 'اطلاعات با ثبت احوال تطابق ندارد. لطفاً کد ملی، نام و تاریخ تولد را دقیق وارد کنید.'
            }, status=400)

        if response.status_code in [200, 201]:
            age = get_age_from_jalali(tarikhe_tavallod)
            is_free = age < 15

            # ===== تخفیف باسا -- باید همینجا (قبل از پرداخت) به کاربر نشون
            # داده بشه، وگرنه قیمتی که سمت مرورگر محاسبه و به درگاه فرستاده
            # می‌شه با قیمت نهایی بلیط (که سمت سرور با تخفیف باسا حساب می‌شه)
            # یکی نیست و کاربر باسایی مبلغ کامل (بدون تخفیف) پرداخت می‌کنه.
            is_basa = False
            basa_discount_percent = 0
            if not is_free and match_id:
                try:
                    match = Match.objects.get(id=match_id)
                    basa_discount_percent = get_basa_discount_percent(match)
                    if basa_discount_percent > 0:
                        is_basa = User.objects.filter(national_code=kode_meli, is_basa_member=True).exists()
                except Match.DoesNotExist:
                    pass

            return JsonResponse({
                'success': True,
                'id': data.get('id', ''),
                'is_free': is_free,
                'age': age,
                'is_basa': is_basa,
                'basa_discount_percent': basa_discount_percent if is_basa else 0,
            })
        else:
            error_msg = data.get('message', data.get('error', 'اطلاعات با ثبت احوال تطابق ندارد.'))
            return JsonResponse({'success': False, 'error': error_msg}, status=response.status_code)

    except Exception as e:
        return JsonResponse({'success': False, 'error': 'خطا در ارتباط با سامانه هواداری. لطفاً مجددا تلاش کنید.'},
                            status=500)


# ============================================================
#  ویوهای مخصوص ادمین (مدیریت و گزارشات)
# ============================================================

@staff_member_required
def sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    match_id = request.GET.get('match_id')

    matches_qs = Match.objects.all().order_by('-date_time')
    if match_id: matches_qs = matches_qs.filter(id=match_id)
    if start_date: matches_qs = matches_qs.filter(date_time__date__gte=start_date)
    if end_date: matches_qs = matches_qs.filter(date_time__date__lte=end_date)

    sales_data = []
    total_sold = total_vip = total_tickets = total_revenue = 0

    for match in matches_qs:
        sold_tickets = Ticket.objects.filter(match=match, status='paid')
        vip_tickets = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued'])

        sold_count = sold_tickets.count()
        vip_count = vip_tickets.count()

        revenue = sum(t.price or 0 for t in sold_tickets)
        vip_revenue = sum(t.price or 0 for t in vip_tickets)

        occupied = MatchSeat.objects.filter(match=match, is_available=False).count()
        total = MatchSeat.objects.filter(match=match).count()

        sales_data.append({
            'match': match, 'sold_count': sold_count, 'vip_count': vip_count,
            'total_tickets': sold_count + vip_count, 'revenue': revenue, 'vip_revenue': vip_revenue,
            'occupancy_percent': round((occupied / total * 100) if total > 0 else 0, 1),
        })

        total_sold += sold_count;
        total_vip += vip_count;
        total_tickets += sold_count + vip_count;
        total_revenue += revenue

    paginator = Paginator(sales_data, 10)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    context = {
        'sales_data': page_obj, 'page_obj': page_obj, 'total_sold': total_sold, 'total_vip': total_vip,
        'total_tickets': total_tickets, 'total_revenue': total_revenue,
        'average_price': total_revenue / total_sold if total_sold > 0 else 0,
        'period_days': (timezone.now().date() - timezone.datetime.strptime(start_date,
                                                                           '%Y-%m-%d').date()).days if start_date else 30,
        'matches': Match.objects.all().order_by('-date_time'), 'selected_match_id': int(match_id) if match_id else None,
        'start_date': start_date, 'end_date': end_date,
    }
    return render(request, 'tickets/sales_report.html', context)


@login_required
@staff_member_required
def bulk_issue_tickets(request):
    if request.method == 'POST':
        form = BulkTicketForm(request.POST)
        if form.is_valid():
            match = form.cleaned_data['match']
            users = form.cleaned_data['users']
            block = form.cleaned_data['block']
            count_per_user = form.cleaned_data['seat_count_per_user']

            seats_in_block = Seat.objects.filter(row__block=block, row__is_active=True, is_available=True).order_by(
                'row__number', 'number')

            available_match_seats = []
            for seat in seats_in_block:
                match_seat, created = MatchSeat.objects.get_or_create(match=match, seat=seat,
                                                                      defaults={'is_available': True})
                if match_seat.is_available: available_match_seats.append(match_seat)

            total_needed = users.count() * count_per_user
            if len(available_match_seats) < total_needed:
                messages.error(request,
                               f'تعداد صندلی‌های موجود در بلوک "{block.name}" ({len(available_match_seats)}) کافی نیست. به {total_needed} صندلی نیاز است.')
                return render(request, 'tickets/bulk_issue.html', {'form': form})

            from matches.models import get_block_price_for_match
            block_price = get_block_price_for_match(match, block) or 0

            with transaction.atomic():
                ticket_index = 0
                created_count = 0
                for user in users:
                    for _ in range(count_per_user):
                        if ticket_index >= len(available_match_seats): break
                        match_seat = available_match_seats[ticket_index]
                        ticket_index += 1
                        if match_seat.match_id != match.id:
                            match_seat.match = match
                            match_seat.save()

                        ticket = Ticket.objects.create(
                            user=user, match=match, seat=match_seat.seat, match_seat=match_seat,
                            full_name=user.get_full_name() or user.username,
                            # Ticket.national_code یک ستون NOT NULL است، ولی کاربران ویژه معمولاً
                            # کد ملی ثبت‌شده ندارند (این فیلد فقط برای ثبت‌نام کاربران عادی
                            # اجباری است) -- getattr به‌تنهایی None را برنمی‌گرداند به رشته‌ی
                            # خالی، چون خودِ فیلد وجود دارد و فقط مقدارش None است؛ همین باعث
                            # خطای IntegrityError/500 هنگام صدور گروهی می‌شد.
                            national_code=getattr(user, 'national_code', '') or '',
                            status='admin_assigned', is_admin_assigned=True, price=block_price,
                        )
                        # .create() قبلاً خودش save() را صدا زده -- یک save() اضافه اینجا چون
                        # pdf_file هنوز خالیه، Ticket.save() را وادار می‌کرد enqueue_pdf_generation
                        # را دوباره صدا بزند و همان بلیط را دوبار در صف PDF قرار بدهد (این دقیقاً
                        # همان چیزی بود که صف تولید PDF را با ده‌ها هزار job تکراری پر کرده بود).
                        match_seat.is_available = False
                        match_seat.save()
                        created_count += 1

                messages.success(request,
                                 f'{created_count} بلیط برای {users.count()} کاربر ویژه از بلوک "{block.name}" صادر شد.')
                return redirect('tickets:bulk_issue')
        else:
            for field, errors in form.errors.items():
                for error in errors: messages.error(request, f'{error}')
    else:
        form = BulkTicketForm()

    return render(request, 'tickets/bulk_issue.html', {'form': form})


@staff_member_required
def manage_vip_users(request):
    match_id = request.GET.get('match_id')
    matches = Match.objects.filter(is_active=True).order_by('-date_time')

    ticket_filter = Q(ticket__is_admin_assigned=True, ticket__status='admin_assigned')
    if match_id:
        ticket_filter &= Q(ticket__match_id=match_id)

    vip_users_qs = User.objects.filter(user_type='vip').annotate(
        ticket_count=Count('ticket', filter=ticket_filter)
    ).order_by('username')

    total_users = vip_users_qs.count()
    total_tickets = vip_users_qs.aggregate(total=Sum('ticket_count'))['total'] or 0
    active_users = vip_users_qs.filter(ticket_count__gt=0).count()
    avg_tickets = total_tickets / total_users if total_users else 0

    paginator = Paginator(vip_users_qs, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj': page_obj, 'matches': matches, 'selected_match_id': int(match_id) if match_id else None,
        'total_users': total_users, 'total_tickets': total_tickets, 'active_users': active_users, 'avg_tickets': avg_tickets,
    }
    return render(request, 'tickets/manage_vip_users.html', context)


@login_required
@staff_member_required
def manage_user_tickets(request, user_id):
    try:
        user = User.objects.get(id=user_id, user_type='vip')
    except User.DoesNotExist:
        messages.error(request, 'کاربر ویژه مورد نظر یافت نشد.')
        return redirect('tickets:manage_vip_users')

    match_id = request.GET.get('match_id')
    tickets = Ticket.objects.filter(user=user, is_admin_assigned=True, status='admin_assigned').select_related('match',
                                                                                                               'seat__row__block').order_by(
        '-purchase_date')

    selected_match = None
    if match_id:
        tickets = tickets.filter(match_id=match_id)
        selected_match = get_object_or_404(Match, id=match_id)

    matches = Match.objects.filter(is_active=True).order_by('-date_time')

    if request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        if ticket_id:
            ticket = get_object_or_404(Ticket, id=ticket_id, user=user)
            with transaction.atomic():
                if ticket.match_seat:
                    ticket.match_seat.is_available = True;
                    ticket.match_seat.save()
                if ticket.pdf_file: ticket.pdf_file.delete(save=False)
                if ticket.qr_code: ticket.qr_code.delete(save=False)
                ticket.delete()
                messages.success(request, f'✅ بلیط {ticket.ticket_number} با موفقیت حذف شد و صندلی آزاد شد.')
            return redirect(
                f'{reverse("tickets:manage_user_tickets", args=[user_id])}?match_id={match_id}' if match_id else reverse(
                    'tickets:manage_user_tickets', args=[user_id]))

    context = {
        'target_user': user, 'tickets': tickets, 'matches': matches,
        'selected_match': selected_match, 'selected_match_id': int(match_id) if match_id else None,
    }
    return render(request, 'tickets/manage_user_tickets.html', context)


@login_required
@staff_member_required
def row_occupancy_report(request):
    matches = Match.objects.all().order_by('-date_time')
    match_id = request.GET.get('match_id')

    if match_id:
        try:
            match_id = int(match_id)
        except (ValueError, TypeError):
            match_id = None
    else:
        match_id = None

    report_data = []
    if match_id:
        match = get_object_or_404(Match, id=match_id)

        # ===== اصلاح مهم: فقط بلوک‌های متعلق به ورزشگاه همین مسابقه =====
        blocks = Block.objects.filter(stadium=match.stadium, is_active=True).order_by('order')

        for block in blocks:
            # ۱. کل صندلی‌های فعال این بلوک
            total_seats = Seat.objects.filter(row__block=block, row__is_active=True).count()

            # ۲. بلیط‌های قطعا فروخته شده برای این مسابقه در این بلوک
            sold_tickets = Ticket.objects.filter(
                match=match,
                seat__row__block=block,
                status__in=['paid', 'admin_assigned', 'vip_issued']
            ).count()

            # ۳. صندلی‌هایی که الان در سبد خرید هستند (رزرو موقت) برای این مسابقه
            reserved_seats = MatchSeat.objects.filter(
                match=match,
                seat__row__block=block,
                is_available=False,
                reserved_until__gt=timezone.now()
            ).count()

            # مجموع اشغال شده‌ها
            occupied = sold_tickets + reserved_seats
            available = max(0, total_seats - occupied)

            report_data.append({
                'row': block, 'total': total_seats, 'occupied': occupied, 'available': available,
                'occupancy_percent': round((occupied / total_seats * 100) if total_seats > 0 else 0, 1)
            })

    # محاسبه مجموع‌ها برای کارت‌های آماری بالای جدول
    total_seats_sum = sum(item['total'] for item in report_data)
    occupied_seats_sum = sum(item['occupied'] for item in report_data)
    available_seats_sum = sum(item['available'] for item in report_data)

    context = {
        'report_data': report_data,
        'matches': matches,
        'selected_match_id': match_id,
        'total_seats_sum': total_seats_sum,
        'occupied_seats_sum': occupied_seats_sum,
        'available_seats_sum': available_seats_sum,
    }
    return render(request, 'tickets/row_occupancy.html', context)

# ============================================================
#  ویوهای مربوط به دانلود گروهی
# ============================================================

@login_required
def bulk_download_tickets(request):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')
    if request.method == 'POST':
        ticket_ids = request.POST.getlist('ticket_ids')
        if not ticket_ids:
            messages.error(request, 'هیچ بلیطی انتخاب نشده است.')
            return redirect('tickets:vip_issued_tickets')
        tickets = Ticket.objects.filter(id__in=ticket_ids, user=request.user, status='vip_issued')
        if not tickets:
            messages.error(request, 'بلیط‌های انتخاب‌شده معتبر نیستند.')
            return redirect('tickets:vip_issued_tickets')
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for ticket in tickets:
                if ticket.pdf_file and ticket.pdf_file.path:
                    try:
                        zip_file.write(ticket.pdf_file.path, arcname=f"ticket_{ticket.ticket_number}.pdf")
                    except FileNotFoundError:
                        pass
        if zip_buffer.getbuffer().nbytes == 0:
            messages.error(request, 'هیچ فایل PDF معتبری برای دانلود وجود ندارد.')
            return redirect('tickets:vip_issued_tickets')
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="tickets.zip"'
        return response
    return redirect('tickets:vip_issued_tickets')


@login_required
def bulk_download_received_tickets(request):
    if request.user.user_type != 'vip':
        messages.error(request, 'شما دسترسی به این بخش ندارید.')
        return redirect('matches:home')
    if request.method == 'POST':
        ticket_ids = request.POST.getlist('ticket_ids')
        if not ticket_ids:
            messages.error(request, 'هیچ بلیطی انتخاب نشده است.')
            return redirect('tickets:vip_tickets')
        tickets = Ticket.objects.filter(id__in=ticket_ids, user=request.user, status='admin_assigned')
        if not tickets:
            messages.error(request, 'بلیط‌های انتخاب‌شده معتبر نیستند.')
            return redirect('tickets:vip_tickets')
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for ticket in tickets:
                if ticket.pdf_file and ticket.pdf_file.path:
                    try:
                        zip_file.write(ticket.pdf_file.path, arcname=f"ticket_{ticket.ticket_number}.pdf")
                    except FileNotFoundError:
                        pass
        if zip_buffer.getbuffer().nbytes == 0:
            messages.error(request, 'هیچ فایل PDF معتبری برای دانلود وجود ندارد.')
            return redirect('tickets:vip_tickets')
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="received_tickets.zip"'
        return response
    return redirect('tickets:vip_tickets')


def check_discount(request):
    if request.method != 'POST': return JsonResponse({'valid': False})
    code = request.POST.get('code', '').strip()
    match_id = request.POST.get('match_id')
    if not match_id: return JsonResponse({'valid': False, 'message': 'شناسه مسابقه مشخص نشده است.'})
    try:
        match = Match.objects.get(id=match_id)
    except Match.DoesNotExist:
        return JsonResponse({'valid': False, 'message': 'مسابقه یافت نشد.'})
    block_id = request.session.get('selected_block_id')
    try:
        discount = DiscountCode.objects.get(code=code)
        block = None
        if block_id: block = Block.objects.get(id=block_id)
        valid, msg = discount.is_valid(match=match, block=block)
        if valid:
            return JsonResponse({'valid': True, 'percent': discount.discount_percent,
                                 'message': f'کد تخفیف {discount.discount_percent}٪ اعمال شد'})
        return JsonResponse({'valid': False, 'message': msg})
    except DiscountCode.DoesNotExist:
        return JsonResponse({'valid': False, 'message': 'کد تخفیف یافت نشد'})


# ============================================================
#  مدیریت تخصیص ظرفیت VIP (فقط ادمین)
# ============================================================

@staff_member_required
def admin_vip_quota_list(request):
    quotas = VIPQuota.objects.select_related('user', 'match').all().order_by('-match__date_time', 'user__username')
    return render(request, 'tickets/admin_vip_quota_list.html', {'quotas': quotas})


@staff_member_required
def admin_vip_quota_create(request):
    if request.method == 'POST':
        form = VIPQuotaForm(request.POST)
        if form.is_valid():
            user = form.cleaned_data['user'];
            match = form.cleaned_data['match']
            if VIPQuota.objects.filter(user=user, match=match).exists():
                messages.error(request, f'کاربر "{user.username}" قبلاً برای این مسابقه ظرفیت دارد.')
                return render(request, 'tickets/admin_vip_quota_form.html', {'form': form, 'is_edit': False})
            form.save()
            messages.success(request, f'ظرفیت برای کاربر "{user.username}" در مسابقه "{match}" با موفقیت ایجاد شد.')
            return redirect('tickets:admin_vip_quota_list')
        else:
            messages.error(request, 'خطا در فرم. لطفاً دوباره تلاش کنید.')
    else:
        form = VIPQuotaForm()
    return render(request, 'tickets/admin_vip_quota_form.html', {'form': form, 'is_edit': False})


@staff_member_required
def admin_vip_quota_edit(request, quota_id):
    quota = get_object_or_404(VIPQuota, id=quota_id)
    if request.method == 'POST':
        form = VIPQuotaForm(request.POST, instance=quota)
        if form.is_valid():
            form.save()
            messages.success(request, f'ظرفیت کاربر "{quota.user.username}" با موفقیت ویرایش شد.')
            return redirect('tickets:admin_vip_quota_list')
        else:
            messages.error(request, 'خطا در فرم. لطفاً دوباره تلاش کنید.')
    else:
        form = VIPQuotaForm(instance=quota)
    return render(request, 'tickets/admin_vip_quota_form.html', {'form': form, 'quota': quota, 'is_edit': True})


@staff_member_required
def admin_vip_quota_delete(request, quota_id):
    quota = get_object_or_404(VIPQuota, id=quota_id)
    if request.method == 'POST':
        user_name = quota.user.username;
        match_name = str(quota.match)
        quota.delete()
        messages.success(request, f'تخصیص ظرفیت کاربر "{user_name}" برای مسابقه "{match_name}" با موفقیت حذف شد.')
        return redirect('tickets:admin_vip_quota_list')
    return render(request, 'tickets/admin_vip_quota_confirm_delete.html', {'quota': quota})


# ============================================================
#  مدیریت کدهای تخفیف (فقط ادمین)
# ============================================================

from django.views.decorators.cache import never_cache


@staff_member_required
@never_cache
def admin_discount_list(request):
    discounts = DiscountCode.objects.select_related('match', 'block').all().order_by('-created_at')
    active_count = discounts.filter(is_active=True).count()
    total_uses = discounts.aggregate(total=Sum('used_count'))['total'] or 0
    avg_discount = discounts.aggregate(avg=Avg('discount_percent'))['avg'] or 0
    context = {
        'discounts': discounts, 'active_count': active_count, 'total_uses': total_uses, 'avg_discount': avg_discount,
    }
    return render(request, 'tickets/admin_discount_list.html', context)


@staff_member_required
def admin_discount_create(request):
    if not Match.objects.filter(is_active=True).exists():
        messages.warning(request, 'برای ایجاد کد تخفیف، ابتدا باید یک مسابقه فعال ایجاد کنید.')
        return redirect('matches:admin_match_create')
    if request.method == 'POST':
        form = DiscountCodeForm(request.POST)
        if form.is_valid():
            discount = form.save()
            messages.success(request, f'کد تخفیف "{discount.code}" با موفقیت ایجاد شد.')
            return redirect('tickets:admin_discount_list')
        else:
            for field, errors in form.errors.items():
                for error in errors: messages.error(request, f'{field}: {error}')
    else:
        form = DiscountCodeForm()
    return render(request, 'tickets/admin_discount_form.html', {'form': form, 'is_edit': False})


@staff_member_required
def admin_discount_edit(request, discount_id):
    discount = get_object_or_404(DiscountCode, id=discount_id)
    if request.method == 'POST':
        form = DiscountCodeForm(request.POST, instance=discount)
        if form.is_valid():
            form.save()
            messages.success(request, f'کد تخفیف "{discount.code}" با موفقیت ویرایش شد.')
            return redirect('tickets:admin_discount_list')
        else:
            messages.error(request, 'خطا در فرم. لطفاً دوباره تلاش کنید.')
    else:
        form = DiscountCodeForm(instance=discount)
    return render(request, 'tickets/admin_discount_form.html', {'form': form, 'discount': discount, 'is_edit': True})


@staff_member_required
def admin_discount_delete(request, discount_id):
    discount = get_object_or_404(DiscountCode, id=discount_id)
    if request.method == 'POST':
        code = discount.code;
        discount.delete()
        messages.success(request, f'کد تخفیف "{code}" با موفقیت حذف شد.')
        return redirect('tickets:admin_discount_list')
    return render(request, 'tickets/admin_discount_confirm_delete.html', {'discount': discount})


@staff_member_required
def admin_discount_toggle(request, discount_id):
    discount = get_object_or_404(DiscountCode, id=discount_id)
    discount.is_active = not discount.is_active;
    discount.save()
    status = 'فعال' if discount.is_active else 'غیرفعال'
    messages.success(request, f'کد تخفیف "{discount.code}" {status} شد.')
    return redirect('tickets:admin_discount_list')


@staff_member_required
def export_sales_report_excel(request):
    if request.method != 'POST':
        messages.error(request, 'درخواست نامعتبر.');
        return redirect('tickets:sales_report')
    match_ids = request.POST.getlist('match_ids')
    if not match_ids:
        messages.warning(request, 'هیچ مسابقه‌ای انتخاب نشده است.');
        return redirect('tickets:sales_report')
    matches = Match.objects.filter(id__in=match_ids).order_by('-date_time')
    if not matches:
        messages.error(request, 'مسابقه‌های انتخاب‌شده یافت نشدند.');
        return redirect('tickets:sales_report')

    wb = Workbook();
    wb.remove(wb.active)
    summary_ws = wb.create_sheet("خلاصه کل", 0)
    summary_ws.append(['گزارش فروش چند مسابقه'])
    summary_ws.merge_cells('A1:F1')
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws['A1'].alignment = Alignment(horizontal='center')
    summary_ws.append([])

    summary_headers = ['ردیف', 'مسابقه', 'رشته', 'تاریخ', 'تعداد فروش', 'درآمد کل (ریال)', 'درصد اشغال']
    summary_ws.append(summary_headers)
    for cell in summary_ws[summary_ws.max_row]:
        cell.font = Font(bold=True);
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")

    total_all = 0;
    revenue_all = 0;
    row_num = 1

    for match in matches:
        sold_tickets = Ticket.objects.filter(match=match, status='paid')
        vip_tickets = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued'])
        sold_count = sold_tickets.count();
        vip_count = vip_tickets.count()
        total_tickets = sold_count + vip_count

        revenue = sum(t.price or 0 for t in sold_tickets)
        vip_revenue = sum(t.price or 0 for t in vip_tickets)
        total_revenue = revenue

        total_seats = Seat.objects.filter(row__block__stadium=match.stadium, row__block__is_active=True,
                                          row__is_active=True).count()
        sold_seats = MatchSeat.objects.filter(match=match, is_available=False).count()
        occupancy = round((sold_seats / total_seats * 100) if total_seats > 0 else 0, 1)

        summary_ws.append([
            row_num, f"{match.home_team} vs {match.away_team}", match.get_sport_type_display(),
            match.date_time.strftime('%Y/%m/%d'), total_tickets, total_revenue, f"{occupancy}%"
        ])
        total_all += total_tickets;
        revenue_all += total_revenue;
        row_num += 1

        sheet_name = f"{match.home_team} - {match.away_team}"[:31]
        ws = wb.create_sheet(sheet_name)

        ws.append([f"گزارش فروش: {match.home_team} vs {match.away_team}"])
        ws.merge_cells('A1:F1');
        ws['A1'].font = Font(bold=True, size=14);
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.append([])
        ws.append(['تاریخ مسابقه:', match.date_time.strftime('%Y/%m/%d %H:%M')])
        ws.append(['ورزشگاه:', match.stadium.name])
        ws.append(['ظرفیت کل:', total_seats])
        ws.append(['تعداد فروش:', total_tickets])
        ws.append(['درآمد کل:', total_revenue])
        ws.append(['درصد اشغال:', f"{occupancy}%"])
        ws.append([])

        headers = ['شماره بلیط', 'کاربر', 'نام خریدار', 'کد ملی', 'قیمت (ریال)', 'وضعیت']
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True);
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")

        all_tickets = list(sold_tickets) + list(vip_tickets)
        for t in all_tickets[:100]:
            ws.append([
                t.ticket_number, t.user.username, t.full_name, t.national_code,
                t.price or 0, t.get_status_display(),
            ])

        for col in ['A', 'B', 'C', 'D', 'E', 'F']: ws.column_dimensions[col].width = 18

    summary_ws.append([])
    summary_ws.append(['جمع کل', '', '', '', total_all, revenue_all, ''])
    for cell in summary_ws[summary_ws.max_row]:
        cell.font = Font(bold=True);
        cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']: summary_ws.column_dimensions[col].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="گزارش_فروش_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def get_blocks_for_match(request):
    match_id = request.GET.get('match_id')
    if not match_id: return JsonResponse({'blocks': []})
    try:
        match = Match.objects.get(id=match_id)
        blocks = Block.objects.filter(stadium=match.stadium, is_active=True).order_by('order')
        blocks_data = [{'id': block.id, 'name': block.name} for block in blocks]
        return JsonResponse({'blocks': blocks_data})
    except Match.DoesNotExist:
        return JsonResponse({'blocks': []})


@login_required
def bulk_download_user_tickets(request):
    if request.method != 'POST':
        messages.error(request, 'درخواست نامعتبر.');
        return redirect('tickets:user_tickets')
    ticket_ids = request.POST.getlist('ticket_ids')
    if not ticket_ids:
        messages.warning(request, 'هیچ بلیطی انتخاب نشده است.');
        return redirect('tickets:user_tickets')
    tickets = Ticket.objects.filter(id__in=ticket_ids, user=request.user,
                                    status__in=['paid', 'admin_assigned', 'vip_issued'])
    if not tickets:
        messages.error(request, 'بلیط‌های انتخاب‌شده معتبر نیستند.');
        return redirect('tickets:user_tickets')
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for ticket in tickets:
            if ticket.pdf_file and ticket.pdf_file.path:
                try:
                    zip_file.write(ticket.pdf_file.path, arcname=f"ticket_{ticket.ticket_number}.pdf")
                except FileNotFoundError:
                    pass
    if zip_buffer.getbuffer().nbytes == 0:
        messages.error(request, 'هیچ فایل PDF معتبری برای دانلود وجود ندارد.');
        return redirect('tickets:user_tickets')
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="tickets_{timezone.now().strftime("%Y%m%d")}.zip"'
    return response
