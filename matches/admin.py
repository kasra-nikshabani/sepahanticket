from django.contrib import admin
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.urls import path, reverse
from django.template.response import TemplateResponse
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ===== ایمپورت مدل‌ها (با اضافه کردن Block) =====
from .models import Match, Row, Seat, Stadium, MatchSeat, Block
from tickets.models import Ticket


def _write_ticket_sheet(wb, sheet_title, report_title, amount_label, qs):
    """یک شیت اکسل با عنوان، خلاصه (مجموع مبلغ + تعداد کل) و جدول بلیط‌ها می‌سازد."""
    ws = wb.create_sheet(sheet_title[:31])
    total_amount = sum(t.price or 0 for t in qs)
    total_count = qs.count()

    ws.append([report_title])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append([amount_label, total_amount, '', 'تعداد کل بلیط:', total_count, ''])
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=3, column=2).font = Font(bold=True, color='006400')
    ws.cell(row=3, column=4).font = Font(bold=True)
    ws.cell(row=3, column=5).font = Font(bold=True, color='006400')
    ws.append([])

    headers = ['شماره بلیط', 'کاربر', 'نام خریدار', 'کد ملی', 'قیمت (ریال)', 'تاریخ']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for t in qs:
        ws.append([
            t.ticket_number, t.user.username, t.full_name, t.national_code,
            t.price or 0,
            t.purchase_date.strftime('%Y/%m/%d %H:%M') if t.purchase_date else '',
        ])

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16

    return total_amount, total_count


def _match_zone_categories(sold_tickets_qs):
    """بلیط‌های فروخته‌شده‌ی یک مسابقه را بر اساس دسته‌بندی جایگاه (میزبان/میهمان/بانوان/کلاس ۱/VIP) تفکیک می‌کند."""
    return [
        ('میزبان', sold_tickets_qs.filter(seat__row__block__zone_type='home')),
        ('میهمان', sold_tickets_qs.filter(seat__row__block__zone_type='away')),
        ('بانوان', sold_tickets_qs.filter(seat__row__block__zone_type='women')),
        ('کلاس ۱', sold_tickets_qs.filter(
            Q(seat__row__block__zone_type='class1') | Q(seat__row__block__is_class1=True))),
        ('VIP', sold_tickets_qs.filter(
            Q(seat__row__block__zone_type='vip') | Q(seat__row__block__is_vip=True))),
    ]


# ============================================================
#  Inline برای نمایش بلیط‌ها (غیرفعال)
# ============================================================
class SoldTicketInline(admin.TabularInline):
    model = Ticket
    extra = 0
    max_num = 0
    show_change_link = False
    can_delete = False
    verbose_name = "بلیط فروخته‌شده"
    verbose_name_plural = "بلیط‌های فروخته‌شده"

    def get_queryset(self, request):
        return Ticket.objects.none()


class VIPAssignedTicketInline(admin.TabularInline):
    model = Ticket
    extra = 0
    max_num = 0
    show_change_link = False
    can_delete = False
    verbose_name = "بلیط تخصیص‌یافته به کاربر ویژه"
    verbose_name_plural = "بلیط‌های تخصیص‌یافته به کاربران ویژه"

    def get_queryset(self, request):
        return Ticket.objects.none()


# ============================================================
#  ادمین مسابقات (با گزارش و اکسل)
# ============================================================
@admin.register(Match)
class MatchAdmin(admin.ModelAdmin):
    change_form_template = 'admin/matches/change_form_with_pagination.html'
    list_display = ('sport_type', 'home_team', 'away_team', 'date_time', 'stadium', 'is_active', 'report_link')
    list_filter = ('sport_type', 'is_active', 'date_time')
    search_fields = ('home_team', 'away_team', 'sport_type')
    inlines = []
    readonly_fields = ('created_by',)
    fields = ('sport_type', 'home_team', 'away_team', 'home_team_logo', 'away_team_logo', 'stadium', 'date_time',
              'is_active', 'created_by')

    def report_link(self, obj):
        url = reverse('admin:match_report', args=[obj.id])
        return format_html('<a href="{}" class="button" target="_blank">📊 گزارش</a>', url)

    report_link.short_description = 'گزارش مسابقه'
    report_link.allow_tags = True

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export/<int:match_id>/', self.export_tickets_view, name='export_tickets'),
            path('export-vip/<int:match_id>/', self.export_vip_tickets_view, name='export_vip_tickets'),
            path('export-categorized/<int:match_id>/', self.export_categorized_tickets_view,
                 name='export_categorized_tickets'),
            path('export-all/<int:match_id>/', self.export_all_tickets_view, name='export_all_tickets'),
            path('report/', self.report_view, name='matches_report'),
            path('report/<int:match_id>/', self.match_report_view, name='match_report'),
        ]
        return custom_urls + urls

    def report_view(self, request):
        matches = Match.objects.all().order_by('-date_time')
        report_data = []
        for match in matches:
            sold_tickets = Ticket.objects.filter(match=match, status='paid')
            vip_tickets = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued'])
            total_tickets = sold_tickets.count() + vip_tickets.count()

            total_revenue = sum(t.price or 0 for t in sold_tickets)
            vip_revenue = sum(t.price or 0 for t in vip_tickets)

            occupied_seats = MatchSeat.objects.filter(match=match, is_available=False).count()
            total_seats = Seat.objects.filter(
                row__block__stadium=match.stadium, row__block__is_active=True, row__is_active=True
            ).count()

            report_data.append({
                'match': match,
                'sold_count': sold_tickets.count(),
                'vip_count': vip_tickets.count(),
                'total_tickets': total_tickets,
                'total_revenue': total_revenue,
                'sold_revenue': total_revenue,
                'vip_revenue': vip_revenue,
                'occupied_seats': occupied_seats,
                'total_seats': total_seats,
                'occupancy_percent': round((occupied_seats / total_seats * 100) if total_seats > 0 else 0, 1),
            })

        context = {
            'report_data': report_data,
            'title': 'گزارش کلی مسابقات',
        }
        return TemplateResponse(request, 'admin/matches/report.html', context)

    def match_report_view(self, request, match_id):
        match = get_object_or_404(Match, id=match_id)
        sold_tickets = Ticket.objects.filter(match=match, status='paid')
        vip_tickets = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued'])

        total_sold = sold_tickets.count()
        total_vip = vip_tickets.count()
        total_tickets = total_sold + total_vip

        sold_revenue = sum(t.price or 0 for t in sold_tickets)
        vip_revenue = sum(t.price or 0 for t in vip_tickets)

        match_seats = MatchSeat.objects.filter(match=match)
        total_match_seats = match_seats.count()
        occupied = match_seats.filter(is_available=False).count()
        available = total_match_seats - occupied

        # ===== دریافت بلوک‌ها (با پشتیبانی از هر دو حالت) =====
        from matches.models import Block

        # روش ۱: استفاده از block_set (پیش‌فرض Django)
        try:
            blocks = match.stadium.block_set.all()
        except AttributeError:
            # روش ۲: استفاده از related_name='blocks'
            if hasattr(match.stadium, 'blocks'):
                blocks = match.stadium.blocks.all()
            else:
                # روش ۳: دریافت مستقیم
                blocks = Block.objects.filter(stadium=match.stadium, is_active=True).order_by('order')

        block_stats = []
        for block in blocks:
            block_seats = MatchSeat.objects.filter(match=match, seat__row__block=block)
            block_total = block_seats.count()
            block_occupied = block_seats.filter(is_available=False).count()
            block_available = block_total - block_occupied
            block_stats.append({
                'block': block,
                'total': block_total,
                'occupied': block_occupied,
                'available': block_available,
                'occupancy': round((block_occupied / block_total * 100) if block_total > 0 else 0, 1),
            })

        context = {
            'match': match,
            'total_sold': total_sold,
            'total_vip': total_vip,
            'total_tickets': total_tickets,
            'sold_revenue': sold_revenue,
            'vip_revenue': vip_revenue,
            'total_revenue': sold_revenue,
            'total_match_seats': total_match_seats,
            'occupied': occupied,
            'available': available,
            'occupancy_percent': round((occupied / total_match_seats * 100) if total_match_seats > 0 else 0, 1),
            'block_stats': block_stats,
            'title': f'گزارش مسابقه {match.home_team} vs {match.away_team}',
        }
        return TemplateResponse(request, 'admin/matches/match_report.html', context)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        match = get_object_or_404(Match, pk=object_id)

        sold_tickets_qs = Ticket.objects.filter(match=match, status='paid').order_by('-purchase_date')
        vip_tickets_qs = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued']).order_by(
            '-purchase_date')

        sold_total = sum(t.price or 0 for t in sold_tickets_qs)
        vip_total = sum(t.price or 0 for t in vip_tickets_qs)

        per_page = 10
        sold_page_num = request.GET.get('sold_page', 1)
        vip_page_num = request.GET.get('vip_page', 1)

        sold_paginator = Paginator(sold_tickets_qs, per_page)
        sold_page = sold_paginator.get_page(sold_page_num)

        vip_paginator = Paginator(vip_tickets_qs, per_page)
        vip_page = vip_paginator.get_page(vip_page_num)

        extra_context.update({
            'match': match,
            'sold_page': sold_page,
            'vip_page': vip_page,
            'sold_total': sold_total,
            'vip_total': vip_total,
            'match_id': object_id,
            'sold_page_range': sold_paginator.page_range,
            'vip_page_range': vip_paginator.page_range,
            'current_sold_page': sold_page.number,
            'current_vip_page': vip_page.number,
        })

        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def export_tickets_view(self, request, match_id):
        match = get_object_or_404(Match, pk=match_id)
        sold_tickets = Ticket.objects.filter(match=match, status='paid')
        vip_tickets = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued'])

        wb = Workbook()
        wb.remove(wb.active)
        _write_ticket_sheet(wb, 'فروخته‌شده', 'گزارش بلیط‌های فروخته‌شده', 'جمع کل فروش:', sold_tickets)
        _write_ticket_sheet(
            wb, 'تخصیص بلیط ویژه', 'گزارش بلیط‌های تخصیص‌یافته به کاربران ویژه', 'جمع کل تخصیص‌ها:', vip_tickets
        )

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="tickets_{match.id}_{match.home_team}_{match.away_team}.xlsx"'
        wb.save(response)
        return response

    def export_vip_tickets_view(self, request, match_id):
        match = get_object_or_404(Match, pk=match_id)
        vip_tickets = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued'])

        wb = Workbook()
        wb.remove(wb.active)
        _write_ticket_sheet(
            wb, 'بلیط‌های ویژه', 'گزارش بلیط‌های تخصیص‌یافته به کاربران ویژه', 'جمع کل تخصیص‌ها:', vip_tickets
        )

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="vip_tickets_{match.id}_{match.home_team}_{match.away_team}.xlsx"'
        wb.save(response)
        return response

    def export_categorized_tickets_view(self, request, match_id):
        """اکسل بلیط‌های فروخته‌شده به تفکیک دسته‌بندی جایگاه (میزبان/میهمان/بانوان/کلاس ۱/VIP)."""
        match = get_object_or_404(Match, pk=match_id)
        sold_tickets = Ticket.objects.filter(match=match, status='paid')

        wb = Workbook()
        wb.remove(wb.active)
        for title, qs in _match_zone_categories(sold_tickets):
            _write_ticket_sheet(wb, title, f'گزارش بلیط‌های فروخته‌شده - {title}', 'جمع مبلغ فروش این بخش:', qs)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f'attachment; filename="tickets_by_category_{match.id}_{match.home_team}_{match.away_team}.xlsx"'
        )
        wb.save(response)
        return response

    def export_all_tickets_view(self, request, match_id):
        """اکسل کامل: همه‌ی دسته‌بندی‌های بلیط فروخته‌شده + بلیط‌های ویژه، هرکدام در شیت جدا با جمع مبلغ و تعداد کل."""
        match = get_object_or_404(Match, pk=match_id)
        sold_tickets = Ticket.objects.filter(match=match, status='paid')
        vip_tickets = Ticket.objects.filter(match=match, status__in=['admin_assigned', 'vip_issued'])

        wb = Workbook()
        wb.remove(wb.active)
        for title, qs in _match_zone_categories(sold_tickets):
            _write_ticket_sheet(wb, title, f'گزارش بلیط‌های فروخته‌شده - {title}', 'جمع مبلغ فروش این بخش:', qs)
        _write_ticket_sheet(
            wb, 'بلیط ویژه', 'گزارش بلیط‌های تخصیص‌یافته به کاربران ویژه', 'جمع کل تخصیص‌ها:', vip_tickets
        )

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f'attachment; filename="tickets_all_{match.id}_{match.home_team}_{match.away_team}.xlsx"'
        )
        wb.save(response)
        return response


# ============================================================
#  ادمین بلوک‌ها (با قابلیت ویرایش قیمت)
# ============================================================
@admin.register(Block)
class BlockAdmin(admin.ModelAdmin):
    list_display = ('name', 'order', 'zone_type', 'price', 'is_vip', 'is_class1', 'is_active')
    list_filter = ('zone_type', 'is_vip', 'is_class1', 'is_active')
    search_fields = ('name',)
    list_editable = ('price', 'order', 'zone_type', 'is_active')
    fields = ('name', 'order', 'zone_type', 'price', 'is_vip', 'is_class1', 'is_active')
    ordering = ('order',)


# ============================================================
#  ادمین سایر مدل‌ها
# ============================================================
@admin.register(Stadium)
class StadiumAdmin(admin.ModelAdmin):
    list_display = ('name', 'capacity')


@admin.register(Row)
class RowAdmin(admin.ModelAdmin):
    list_display = ('block', 'number', 'is_active')
    list_filter = ('block', 'is_active')
    search_fields = ('block__name',)


@admin.register(Seat)
class SeatAdmin(admin.ModelAdmin):
    list_display = ('row', 'number', 'is_available')
    list_filter = ('row__block', 'is_available')
    search_fields = ('row__block__name',)
    raw_id_fields = ('row',)


@admin.register(MatchSeat)
class MatchSeatAdmin(admin.ModelAdmin):
    list_display = ('match', 'seat', 'is_available', 'reserved_until')
    list_filter = ('match', 'is_available')
    search_fields = ('match__home_team', 'seat__row__block__name')
    raw_id_fields = ('match', 'seat')


# matches/admin.py
from django.contrib import admin
from django.utils.html import format_html
from .models import Match, MatchCost, MatchRevenue, MatchFinancialReport


@admin.register(MatchCost)
class MatchCostAdmin(admin.ModelAdmin):
    list_display = ['match', 'description', 'amount_display', 'created_at']
    list_filter = ['match', 'created_at']
    search_fields = ['match__home_team', 'match__away_team', 'description']
    readonly_fields = ['created_at', 'updated_at']

    def amount_display(self, obj):
        return f"{obj.amount:,} ریال"

    amount_display.short_description = 'مبلغ'


@admin.register(MatchRevenue)
class MatchRevenueAdmin(admin.ModelAdmin):
    list_display = ['match', 'description', 'amount_display', 'created_at']
    list_filter = ['match', 'created_at']
    search_fields = ['match__home_team', 'match__away_team', 'description']
    readonly_fields = ['created_at', 'updated_at']

    def amount_display(self, obj):
        return f"{obj.amount:,} ریال"

    amount_display.short_description = 'مبلغ'


@admin.register(MatchFinancialReport)
class MatchFinancialReportAdmin(admin.ModelAdmin):
    list_display = [
        'match',
        'total_ticket_sold',
        'total_vip_tickets',
        'total_used_tickets',
        'total_ticket_revenue_display',
        'total_wallet_usage_display',  # ← جدید
        'total_costs_display',
        'total_revenues_display',
        'net_profit_display',
        'generated_at',
    ]
    list_filter = ['match', 'generated_at']
    search_fields = ['match__home_team', 'match__away_team']
    readonly_fields = [
        'match',
        'total_ticket_revenue',
        'total_ticket_sold',
        'total_vip_tickets',
        'total_used_tickets',
        'total_costs',
        'total_revenues',
        'net_profit',
        'generated_at',
        'updated_at',
    ]

    def total_wallet_usage_display(self, obj):
        return f"{obj.total_wallet_usage:,} ریال"

    total_wallet_usage_display.short_description = 'مصرف از کیف پول'

    def total_ticket_revenue_display(self, obj):
        return f"{obj.total_ticket_revenue:,} ریال"

    total_ticket_revenue_display.short_description = 'درآمد فروش بلیط'

    def total_costs_display(self, obj):
        return f"{obj.total_costs:,} ریال"

    total_costs_display.short_description = 'مجموع هزینه‌ها'

    def total_revenues_display(self, obj):
        return f"{obj.total_revenues:,} ریال"

    total_revenues_display.short_description = 'درآمدهای اضافی'

    def net_profit_display(self, obj):
        color = 'green' if obj.net_profit >= 0 else 'red'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} ریال</span>',
            color,
            f"{obj.net_profit:,}"
        )

    net_profit_display.short_description = 'سود خالص'

    def has_add_permission(self, request):
        return False  # جلوگیری از ایجاد دستی گزارش

    def has_delete_permission(self, request, obj=None):
        return False  # جلوگیری از حذف گزارش
